import openpyxl
import re
import telebot
from telebot import *
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton
from telebot import types
import os
import threading
import time
from datetime import datetime
import json

# Маппинг дней недели и ячеек для времени и пар
DAYS_MAPPING = {
    "понедельник": {
        "pairs_cells": ["A9", "A11", "A13", "A15", "A17", "A19", "A21"],
        "time_cells": [("B9", "B10"), ("B11", "B12"), ("B13", "B14"), ("B15", "B16"),
                       ("B17", "B18"), ("B19", "B20"), ("B21", "B22")],
        "date_cell": "E8"  # Ячейка с датой для SPO
    },
    "вторник": {
        "pairs_cells": ["A24", "A26", "A28", "A30", "A32", "A34", "A36"],
        "time_cells": [("B24", "B25"), ("B26", "B27"), ("B28", "B29"), ("B30", "B31"),
                       ("B32", "B33"), ("B34", "B35"), ("B36", "B37")],
        "date_cell": "E23"  # Ячейка с датой для SPO
    },
    "среда": {
        "pairs_cells": ["A39", "A41", "A43", "A45", "A47", "A49", "A51"],
        "time_cells": [("B39", "B40"), ("B41", "B42"), ("B43", "B44"), ("B45", "B46"),
                       ("B47", "B48"), ("B49", "B50"), ("B51", "B52")],
        "date_cell": "E38"  # Ячейка с датой для SPO
    },
    "четверг": {
        "pairs_cells": ["A54", "A56", "A58", "A60", "A62", "A64", "A66"],
        "time_cells": [("B54", "B55"), ("B56", "B57"), ("B58", "B59"), ("B60", "B61"),
                       ("B62", "B63"), ("B64", "B65"), ("B66", "B67")],
        "date_cell": "E53"  # Ячейка с датой для SPO
    },
    "пятница": {
        "pairs_cells": ["A69", "A71", "A73", "A75", "A77", "A79", "A81"],
        "time_cells": [("B69", "B70"), ("B71", "B72"), ("B73", "B74"), ("B75", "B76"),
                       ("B77", "B78"), ("B79", "B80"), ("B81", "B82")],
        "date_cell": "E68"  # Ячейка с датой для SPO
    },
    "суббота": {
        "pairs_cells": ["A84", "A86", "A88", "A90", "A92"],
        "time_cells": [("B84", "B85"), ("B86", "B87"), ("B88", "B89"), ("B90", "B91"),
                       ("B92", "B93")],
        "date_cell": "E83"  # Ячейка с датой для SPO
    }
}

VO_DAYS_MAPPING = {
    "лист1": {
        "понедельник": {
            "pairs_cells": ["A9", "A11", "A13", "A15", "A17", "A19", "A21"],
            "time_cells": [("B9", "B10"), ("B11", "B12"), ("B13", "B14"), ("B15", "B16"),
                           ("B17", "B18"), ("B19", "B20"), ("B21", "B22")],
            "date_cell": "E8"
        },
        "вторник": {
            "pairs_cells": ["A24", "A26", "A28", "A30", "A32", "A34", "A36"],
            "time_cells": [("B24", "B25"), ("B26", "B27"), ("B28", "B29"), ("B30", "B31"),
                           ("B32", "B33"), ("B34", "B35"), ("B36", "B37")],
            "date_cell": "E23"
        },
        "среда": {
            "pairs_cells": ["A39", "A41", "A43", "A45", "A47", "A49", "A51"],
            "time_cells": [("B39", "B40"), ("B41", "B42"), ("B43", "B44"), ("B45", "B46"),
                           ("B47", "B48"), ("B49", "B50"), ("B51", "B52")],
            "date_cell": "E38"
        },
        "четверг": {
            "pairs_cells": ["A54", "A56", "A58", "A60", "A62", "A64", "A66"],
            "time_cells": [("B54", "B55"), ("B56", "B57"), ("B58", "B59"), ("B60", "B61"),
                           ("B62", "B63"), ("B64", "B65"), ("B66", "B67")],
            "date_cell": "E53"
        },
        "пятница": {
            "pairs_cells": ["A69", "A71", "A73", "A75", "A77", "A79", "A81"],
            "time_cells": [("B69", "B70"), ("B71", "B72"), ("B73", "B74"), ("B75", "B76"),
                           ("B77", "B78"), ("B79", "B80"), ("B81", "B82")],
            "date_cell": "E68"
        },
        "суббота": {
            "pairs_cells": ["A84", "A86", "A88", "A90", "A92"],
            "time_cells": [("B84", "B85"), ("B86", "B87"), ("B88", "B89"), ("B90", "B91"),
                           ("B92", "B93")],
            "date_cell": "E83"
        }
    },
    "лист2": {
        "понедельник": {
            "pairs_cells": ["A9", "A11", "A13", "A15", "A17", "A19", "A21"],
            "time_cells": [("B9", "B10"), ("B11", "B12"), ("B13", "B14"), ("B15", "B16"),
                           ("B17", "B18"), ("B19", "B20"), ("B21", "B22")],
            "date_cell": "E8"
        },
        "вторник": {
            "pairs_cells": ["A24", "A26", "A28", "A30", "A32", "A34", "A36"],
            "time_cells": [("B24", "B25"), ("B26", "B27"), ("B28", "B29"), ("B30", "B31"),
                           ("B32", "B33"), ("B34", "B35"), ("B36", "B37")],
            "date_cell": "E23"
        },
        "среда": {
            "pairs_cells": ["A39", "A41", "A43", "A45", "A47", "A49", "A51"],
            "time_cells": [("B39", "B40"), ("B41", "B42"), ("B43", "B44"), ("B45", "B46"),
                           ("B47", "B48"), ("B49", "B50"), ("B51", "B52")],
            "date_cell": "E38"
        },
        "четверг": {
            "pairs_cells": ["A54", "A56", "A58", "A60", "A62", "A64", "A66"],
            "time_cells": [("B54", "B55"), ("B56", "B57"), ("B58", "B59"), ("B60", "B61"),
                           ("B62", "B63"), ("B64", "B65"), ("B66", "B67")],
            "date_cell": "E53"
        },
        "пятница": {
            "pairs_cells": ["A69", "A71", "A73", "A75", "A77", "A79", "A81"],
            "time_cells": [("B69", "B70"), ("B71", "B72"), ("B73", "B74"), ("B75", "B76"),
                           ("B77", "B78"), ("B79", "B80"), ("B81", "B82")],
            "date_cell": "E68"
        },
        "суббота": {
            "pairs_cells": ["A84", "A86", "A88", "A90", "A92"],
            "time_cells": [("B84", "B85"), ("B86", "B87"), ("B88", "B89"), ("B90", "B91"),
                           ("B92", "B93")],
            "date_cell": "E83"
        }
    },
    "лист3": {
        "понедельник": {
            "pairs_cells": ["A9", "A11", "A13", "A15", "A17", "A19", "A21"],
            "time_cells": [("B9", "B10"), ("B11", "B12"), ("B13", "B14"), ("B15", "B16"),
                           ("B17", "B18"), ("B19", "B20"), ("B21", "B22")],
            "date_cell": "E8"
        },
        "вторник": {
            "pairs_cells": ["A24", "A26", "A28", "A30", "A32", "A34", "A36"],
            "time_cells": [("B24", "B25"), ("B26", "B27"), ("B28", "B29"), ("B30", "B31"),
                           ("B32", "B33"), ("B34", "B35"), ("B36", "B37")],
            "date_cell": "E24"
        },
        "среда": {
            "pairs_cells": ["A39", "A41", "A43", "A45", "A47", "A49", "A51"],
            "time_cells": [("B39", "B40"), ("B41", "B42"), ("B43", "B44"), ("B45", "B46"),
                           ("B47", "B48"), ("B49", "B50"), ("B51", "B52")],
            "date_cell": "E38"
        },
        "четверг": {
            "pairs_cells": ["A54", "A56", "A58", "A60", "A62", "A64", "A66"],
            "time_cells": [("B54", "B55"), ("B56", "B57"), ("B58", "B59"), ("B60", "B61"),
                           ("B62", "B63"), ("B64", "B65"), ("B66", "B67")],
            "date_cell": "E53"
        },
        "пятница": {
            "pairs_cells": ["A69", "A71", "A73", "A75", "A77", "A79", "A81"],
            "time_cells": [("B69", "B70"), ("B71", "B72"), ("B73", "B74"), ("B75", "B76"),
                           ("B77", "B78"), ("B79", "B80"), ("B81", "B82")],
            "date_cell": "E68"
        },
        "суббота": {
            "pairs_cells": ["A84", "A86", "A88", "A90", "A92"],
            "time_cells": [("B84", "B85"), ("B86", "B87"), ("B88", "B89"), ("B90", "B91"),
                           ("B92", "B93")],
            "date_cell": "E83"
        }
    },
    "лист4": {
        "понедельник": {
            "pairs_cells": ["A9", "A11", "A13", "A15", "A17", "A19", "A21"],
            "time_cells": [("B9", "B10"), ("B11", "B12"), ("B13", "B14"), ("B15", "B16"),
                           ("B17", "B18"), ("B19", "B20"), ("B21", "B22")],
            "date_cell": "E8"
        },
        "вторник": {
            "pairs_cells": ["A24", "A26", "A28", "A30", "A32", "A34", "A36"],
            "time_cells": [("B24", "B25"), ("B26", "B27"), ("B28", "B29"), ("B30", "B31"),
                           ("B32", "B33"), ("B34", "B35"), ("B36", "B37")],
            "date_cell": "E24"
        },
        "среда": {
            "pairs_cells": ["A39", "A41", "A43", "A45", "A47", "A49", "A51"],
            "time_cells": [("B39", "B40"), ("B41", "B42"), ("B43", "B44"), ("B45", "B46"),
                           ("B47", "B48"), ("B49", "B50"), ("B51", "B52")],
            "date_cell": "E38"
        },
        "четверг": {
            "pairs_cells": ["A54", "A56", "A58", "A60", "A62", "A64", "A66"],
            "time_cells": [("B54", "B55"), ("B56", "B57"), ("B58", "B59"), ("B60", "B61"),
                           ("B62", "B63"), ("B64", "B65"), ("B66", "B67")],
            "date_cell": "E53"
        },
        "пятница": {
            "pairs_cells": ["A69", "A71", "A73", "A75", "A77", "A79", "A81"],
            "time_cells": [("B69", "B70"), ("B71", "B72"), ("B73", "B74"), ("B75", "B76"),
                           ("B77", "B78"), ("B79", "B80"), ("B81", "B82")],
            "date_cell": "E68"
        },
        "суббота": {
            "pairs_cells": ["A84", "A86", "A88", "A90", "A92"],
            "time_cells": [("B84", "B85"), ("B86", "B87"), ("B88", "B89"), ("B90", "B91"),
                           ("B92", "B93")],
            "date_cell": "E83"
        }
    }
}

GROUP_ROOM_MAPPING = {
    "5-8124": {"group_cell": "C7", "room_column": "F"},
    "5-8123, 5-2123": {"group_cell": "C7", "room_column": "E"},
    "5-8122, 5-2122": {"group_cell": "B5", "room_column": "D"},
    "5-8121, 5-2121": {"group_cell": "C7", "room_column": "F"}
}

token = "7572602237:AAH4hzckLH6JrqmkSJUBYGG0h-6qgLf1s-k"
bot = telebot.TeleBot(token)

# Временное хранилище для данных пользователя
user_data = {}

def get_group_and_room_cells(sheet):
    """
    Возвращает список ячеек с группами (если заполнены) и колонку аудиторий.
    """
    group_cells = []
    
    # Проверяем, есть ли данные в C7
    if sheet["C7"].value:
        group_cells.append("C7")

    # Проверяем, есть ли данные в E7
    if sheet["E7"].value:
        group_cells.append("E7")

    # Если нет данных в обеих ячейках, используем только C7
    if not group_cells:
        group_cells = ["C7"]

    return {"group_cells": group_cells, "room_column": "D"}



def get_group_data(sheet, group_name):
    """
    Найти группу на листе и вернуть столбец для расписания.
    """
    group_cell_mapping = {"C7": "C", "E7": "E"}  # Группы на листе
    for cell, column in group_cell_mapping.items():
        cell_value = sheet[cell].value
        if cell_value and group_name in cell_value:
            return column
    return None

def determine_vo_sheet_mapping(sheet_name):
    normalized_sheet_name = sheet_name.strip().lower()

    if "лист1" in normalized_sheet_name:
        return VO_DAYS_MAPPING["лист1"]
    elif "лист2" in normalized_sheet_name:
        return VO_DAYS_MAPPING["лист2"]
    elif "лист3" in normalized_sheet_name:
        return VO_DAYS_MAPPING["лист3"]
    elif "лист4" in normalized_sheet_name:
        return VO_DAYS_MAPPING["лист4"]

    # Обработка листов с цифровыми названиями
    if re.match(r"^\d+-\d+", normalized_sheet_name):
        return VO_DAYS_MAPPING.get("лист1")
    return None


def normalize_teacher_name(teacher_name):
    if teacher_name:
        teacher_name = re.sub(r"(преп\.|ст\.преп\.|куратор|к\.э\.н\.|к\.п\.н\.|доцент)\s*", "", teacher_name)
        teacher_name = re.sub(r"\s+", " ", teacher_name).strip().lower()
        return re.sub(r"\.\s+", ".", teacher_name)
    return None

def clean_text(value):
    if value:
        return " ".join(str(value).split())
    return None

def display_schedule(schedule, entity):
    if not schedule:
        print(f"Расписание для {entity} отсутствует.")
        return
    result = []
    for entry in schedule:
        result.append(
            f"\n📅 День недели: {entry.get('day', 'Не указано')}\n"
            f"📅 Дата: {entry['date']}\n"
            f"№ Пары: {entry['pair']}\n"
            f"🕒 Время: {entry['time'][0]} - {entry['time'][1]}\n"
            f"📖 Предмет: {entry['subject']}\n"
            f"🏫 Аудитория: {entry['room']}"
        )
        if 'teacher' in entry:
            result.append(f"👩‍🏫 Преподаватель: {entry['teacher']}")
        if 'group' in entry:
            result.append(f"👨‍🎓 Группа: {entry['group']}")


    return "\n\n".join(result)


def search_teacher(sheet, sheet_name, teacher_name, target_day=None, education_type="SPO"):
    """
    Ищет расписание преподавателя по всем дням недели.
    Если target_day передан, то ищет только для этого дня.
    """
    teacher_name = normalize_teacher_name(teacher_name)
    result = []

    # Получение маппинга для дней недели
    if education_type == "SPO":
        days_mapping = DAYS_MAPPING
    else:
        vo_mapping = determine_vo_sheet_mapping(sheet_name)
        if vo_mapping is None:
            return []
        days_mapping = vo_mapping

    # Получаем все возможные столбцы групп (например, ["C7", "E7"])
    group_cells = get_group_and_room_cells(sheet)["group_cells"]
    columns = [cell[0] for cell in group_cells]  # Преобразуем ["C7", "E7"] в ["C", "E"]

    for day in days_mapping.keys():
        if target_day and target_day.lower() != day.lower():
            continue

        day_info = days_mapping[day]
        pairs_cells = day_info["pairs_cells"]
        time_cells = day_info["time_cells"]
        date_cell = day_info["date_cell"]

        # Извлекаем дату
        date = sheet[date_cell].value

        for column in columns:  # Перебираем "C" и "E"
            room_column = "D" if column == "C" else "F"

            for pair_cell, (time_start, time_end) in zip(pairs_cells, time_cells):
                pair_number = clean_text(sheet[f"A{pair_cell[1:]}"].value)
                subject = clean_text(sheet[f"{column}{pair_cell[1:]}"].value)
                teacher = clean_text(sheet[f"{column}{int(pair_cell[1:]) + 1}"].value)
                room = clean_text(sheet[f"{room_column}{pair_cell[1:]}"].value)

                # Проверяем совпадение имени преподавателя
                if teacher and teacher_name in normalize_teacher_name(teacher):
                    group_cell = f"{column}7"  # Определяем ячейку с группой (C7 или E7)
                    group_name = clean_text(sheet[group_cell].value)  # Получаем название группы

                    result.append({
                        "day": day.capitalize(),
                        "date": date or "Не указана",
                        "time": (sheet[time_start].value, sheet[time_end].value),
                        "subject": subject or "Не указано",
                        "room": room or "Не указана",
                        "group": group_name or "Не указана",  # Теперь у пары правильная группа
                        "pair": pair_number or "Не указано"
                    })

    return result

def extract_schedule(sheet, column, day, education_type, sheet_name):
    schedule = []

    if education_type == "SPO":
        day_info = DAYS_MAPPING[day.lower()]
    else:
        vo_mapping = determine_vo_sheet_mapping(sheet_name)
        if not vo_mapping:
            return []
        day_info = vo_mapping[day.lower()]

    pairs_cells = day_info["pairs_cells"]
    time_cells = day_info["time_cells"]
    room_column = "D" if column == "C" else "F"
    date_cell = day_info["date_cell"]
    date = sheet[date_cell].value

    for pair_cell, (time_start, time_end) in zip(pairs_cells, time_cells):
        pair_number = clean_text(sheet[f"A{pair_cell[1:]}"].value)
        subject = clean_text(sheet[f"{column}{pair_cell[1:]}"].value)
        teacher = clean_text(sheet[f"{column}{int(pair_cell[1:]) + 1}"].value)
        room = clean_text(sheet[f"{room_column}{pair_cell[1:]}"].value)

        if subject or teacher or room:
            schedule.append({
                "day": day.capitalize(),
                "pair": pair_number or "Не указано",
                "time": (sheet[time_start].value, sheet[time_end].value),
                "subject": subject or "Не указано",
                "teacher": teacher or "Не указан",
                "room": room or "Не указана",
                "date": date or "Не указана"
            })
    return schedule

user_data = {}
def back_button(callback_data):
    """Создает кнопку 'Назад' для возврата на предыдущий этап."""
    return types.InlineKeyboardButton("🔙 Назад", callback_data=callback_data)

# Обработчик команды /start
@bot.message_handler(commands=['start', 'help'])
def send_welcome(message):
    chat_id = message.chat.id
    
    # Добавляем пользователя в список, если его нет
    if str(chat_id) not in users:
        users[str(chat_id)] = True
        save_users(users)
    markup = types.InlineKeyboardMarkup()
    markup.add(
        types.InlineKeyboardButton("Студент", callback_data="search_student"),
        types.InlineKeyboardButton("Преподаватель", callback_data="search_teacher")
    )
    bot.send_message(message.chat.id, "Выберите, кого искать:", reply_markup=markup)

# Обработчик выбора "Студент"
@bot.callback_query_handler(func=lambda call: call.data == "search_student")
def select_education_type(call):
    user_data[call.message.chat.id] = {"search_type": "student"}
    
    # Новый текст для сообщения
    new_text = "Выберите тип образования:"
    
    # Создание клавиатуры с кнопкой "Назад"
    markup = types.InlineKeyboardMarkup()
    markup.add(
        types.InlineKeyboardButton("SPO", callback_data="education_spo"),
        types.InlineKeyboardButton("VO", callback_data="education_vo")
    )
    markup.add(back_button("back_to_search"))  # Кнопка "Назад" на предыдущий этап
    
    # Удаление текущего сообщения и отправка нового
    try:
        bot.delete_message(call.message.chat.id, call.message.message_id)
    except Exception as e:
        print(f"Не удалось удалить сообщение: {e}")

    # Отправка нового сообщения
    bot.send_message(call.message.chat.id, new_text, reply_markup=markup)


# Обработчик выбора недели
@bot.callback_query_handler(func=lambda call: call.data.startswith("week_"))
def select_week(call):
    week = call.data.split("_")[1]
    user_data[call.message.chat.id]["week"] = week  # Сохраняем выбранную неделю

    # Удаляем старое сообщение
    try:
        bot.delete_message(call.message.chat.id, call.message.message_id)
    except Exception as e:
        print(f"Не удалось удалить сообщение: {e}")

    # Переходим к этапу выбора дня недели
    markup = types.InlineKeyboardMarkup()
    for day in DAYS_MAPPING.keys():
        markup.add(types.InlineKeyboardButton(day.capitalize(), callback_data=f"day_{day}"))
    markup.add(back_button("education_spo"))  # Кнопка "Назад" на предыдущий этап
    bot.send_message(call.message.chat.id, "Выберите день недели:", reply_markup=markup)


# Модификация функции выбора дня недели
def select_day(message):
    user_data[message.chat.id]["group_name"] = message.text.strip()

    # Удаляем сообщение пользователя
    try:
        bot.delete_message(message.chat.id, message.message_id)
    except Exception as e:
        print(f"Не удалось удалить сообщение пользователя: {e}")

    # Удаляем предыдущее сообщение бота
    last_bot_message_id = user_data[message.chat.id].get("last_bot_message_id")
    if last_bot_message_id:
        try:
            bot.delete_message(message.chat.id, last_bot_message_id)
        except Exception as e:
            print(f"Не удалось удалить сообщение бота: {e}")

    # Отправляем сообщение с выбором недели
    markup = types.InlineKeyboardMarkup()
    markup.add(
        types.InlineKeyboardButton("Эта неделя", callback_data="week_this"),
        types.InlineKeyboardButton("Следующая неделя", callback_data="week_next")
    )
    markup.add(back_button("education_spo"))  # Кнопка "Назад" на выбор типа образования
    bot.send_message(message.chat.id, "Выберите неделю:", reply_markup=markup)


# Обработчик выбора типа образования для студента
@bot.callback_query_handler(func=lambda call: call.data.startswith("education_"))
def ask_group_name(call):
    education_type = call.data.split("_")[1].upper()
    user_data[call.message.chat.id]["education_type"] = education_type

    # Удаляем старое сообщение
    try:
        bot.delete_message(call.message.chat.id, call.message.message_id)
    except Exception as e:
        print(f"Не удалось удалить сообщение: {e}")

    # Отправляем новое сообщение
    markup = types.InlineKeyboardMarkup()
    markup.add(back_button("search_student"))  # Кнопка "Назад"
    sent_message = bot.send_message(call.message.chat.id, "Введите номер группы(например C7124Б. C - eng, Б - ru):", reply_markup=markup)
    
    # Сохраняем ID отправленного сообщения для последующего удаления
    user_data[call.message.chat.id]["last_bot_message_id"] = sent_message.message_id
    bot.register_next_step_handler(sent_message, select_day)



# Обработчик выбора дня недели для студента
@bot.callback_query_handler(func=lambda call: call.data.startswith("day_"))
def show_schedule(call):
    day = call.data.split("_")[1]
    user_data[call.message.chat.id]["day"] = day
    week = user_data[call.message.chat.id].get("week")
    education_type = user_data[call.message.chat.id]["education_type"]
    group_name = user_data[call.message.chat.id]["group_name"]

    # Определяем файл в зависимости от недели и типа образования
    if week == "this":
        file_name = "this_spo.xlsx" if education_type == "SPO" else "this_vo.xlsx"
    elif week == "next":
        file_name = "next_spo.xlsx" if education_type == "SPO" else "next_vo.xlsx"

    try:
        workbook = openpyxl.load_workbook(file_name, data_only=True)
    except FileNotFoundError:
        bot.send_message(call.message.chat.id, f"Ошибка: Файл {file_name} не найден.")
        return

    # Оставшаяся часть логики аналогична
    found_group = False
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        column = get_group_data(sheet, group_name)
        if column:
            schedule = extract_schedule(sheet, column, day, education_type, sheet_name)
            
            # Удаление старого сообщения
            try:
                bot.delete_message(call.message.chat.id, call.message.message_id)
            except Exception as e:
                print(f"Не удалось удалить сообщение: {e}")

            # Если расписание пустое, отправляем сообщение об отсутствии пар
            if not schedule:
                markup = types.InlineKeyboardMarkup()
                markup.add(back_button("back_to_day_selection"))  # Кнопка "Назад" на выбор дня недели
                bot.send_message(call.message.chat.id, f"На {day.capitalize()} для группы {group_name} пар нет.", reply_markup=markup)

            else:
                markup = types.InlineKeyboardMarkup()
                markup.add(back_button("back_to_day_selection"))  # Кнопка возвращает на выбор дня недели
                bot.send_message(call.message.chat.id, display_schedule(schedule, group_name), reply_markup=markup)

            found_group = True
            break

    if not found_group:
        bot.send_message(call.message.chat.id, f"Группа '{group_name}' не найдена в файле {file_name}.")


# Обработчик выбора "Преподаватель"
@bot.callback_query_handler(func=lambda call: call.data == "search_teacher")
def ask_teacher_name(call):
    """ Запрос имени преподавателя """
    user_data[call.message.chat.id] = {"search_type": "teacher"}

    # Удаляем старое сообщение
    try:
        bot.delete_message(call.message.chat.id, call.message.message_id)
    except Exception as e:
        print(f"Ошибка удаления сообщения: {e}")

    # Клавиатура с кнопкой "Назад"
    markup = types.InlineKeyboardMarkup()
    markup.add(back_button("back_to_search"))

    sent_message = bot.send_message(
        call.message.chat.id,
        "Введите фамилию и инициалы преподавателя (например, Иванов И. И.):",
        reply_markup=markup
    )

    user_data[call.message.chat.id]["last_bot_message_id"] = sent_message.message_id
    bot.register_next_step_handler(sent_message, select_teacher_week_step)


def select_teacher_week_step(message):
    """ Запрос выбора недели после ввода преподавателя """
    user_data[message.chat.id]["teacher_name"] = message.text.strip()

    # Удаление сообщения пользователя
    try:
        bot.delete_message(message.chat.id, message.message_id)
    except Exception as e:
        print(f"Ошибка удаления сообщения: {e}")

    # Выбор недели
    markup = types.InlineKeyboardMarkup()
    markup.add(
        types.InlineKeyboardButton("Эта неделя", callback_data="teacher_week_this"),
        types.InlineKeyboardButton("Следующая неделя", callback_data="teacher_week_next")
    )
    markup.add(back_button("search_teacher"))

    bot.send_message(message.chat.id, "Выберите неделю:", reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data.startswith("teacher_week_"))
def select_teacher_week(call):
    """ Обработчик выбора недели преподавателя """
    week = call.data.split("_")[2]
    user_data[call.message.chat.id]["week"] = week

    print(f"[DEBUG] Выбрана неделя: {week}")

    # Выбор дня недели
    markup = types.InlineKeyboardMarkup()
    for day in DAYS_MAPPING.keys():
        markup.add(types.InlineKeyboardButton(day.capitalize(), callback_data=f"teacher_day_{day}"))

    markup.add(back_button("search_teacher"))
    bot.send_message(call.message.chat.id, "Выберите день недели:", reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data.startswith("teacher_day_"))
def show_teacher_schedule(call):
    """ Показывает расписание преподавателя """
    day = call.data.split("_")[2]
    chat_id = call.message.chat.id
    teacher_name = user_data[chat_id]["teacher_name"]
    week = user_data[chat_id].get("week")

    # Файлы для текущей или следующей недели
    file_names = ["this_spo.xlsx", "this_vo.xlsx"] if week == "this" else ["next_spo.xlsx", "next_vo.xlsx"]

    print(f"[DEBUG] Поиск для {teacher_name}, день: {day}, неделя: {week}")

    found_data = []

    for file_name in file_names:
        try:
            workbook = openpyxl.load_workbook(file_name, data_only=True)
        except FileNotFoundError:
            bot.send_message(chat_id, f"Ошибка: Файл {file_name} не найден.")
            continue

        education_type = "SPO" if "spo" in file_name.lower() else "VO"

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Для VO проверяем, есть ли маппинг
            if education_type == "VO":
                vo_mapping = determine_vo_sheet_mapping(sheet_name)
                if not vo_mapping:
                    continue

            # Определение столбца (например, "C" для SPO)
            group_cells = get_group_and_room_cells(sheet)["group_cells"]  # Получаем список групповых ячеек
            columns = [cell[0] for cell in group_cells]  # Преобразуем ["C7", "E7"] в ["C", "E"]



            # Поиск расписания преподавателя
            results = search_teacher(sheet, sheet_name, teacher_name, target_day=day, education_type=education_type)
            found_data.extend(results)

    # Удаление старого сообщения
    try:
        bot.delete_message(chat_id, call.message.message_id)
    except Exception as e:
        print(f"Ошибка удаления сообщения: {e}")

    # Вывод результата
    markup = types.InlineKeyboardMarkup()
    markup.add(back_button(f"teacher_week_{week}"))

    if not found_data:
        bot.send_message(chat_id, f"На {day.capitalize()} для преподавателя {teacher_name} пар нет.", reply_markup=markup)
    else:
        bot.send_message(chat_id, display_schedule(found_data, teacher_name), reply_markup=markup)




# Обработчик кнопки "Назад"
@bot.callback_query_handler(func=lambda call: call.data.startswith("back_"))
def go_back(call):
    if call.data == "back_to_search":
        # Возврат к выбору "Студент/Преподаватель"
        markup = types.InlineKeyboardMarkup()
        markup.add(
            types.InlineKeyboardButton("Студент", callback_data="search_student"),
            types.InlineKeyboardButton("Преподаватель", callback_data="search_teacher")
        )
        
        bot.edit_message_text("Выберите, кого искать:", call.message.chat.id, call.message.message_id, reply_markup=markup)

    elif call.data == "back_to_group_selection":
        # Возврат на выбор типа образования
        markup = types.InlineKeyboardMarkup()
        markup.add(
            types.InlineKeyboardButton("SPO", callback_data="education_spo"),
            types.InlineKeyboardButton("VO", callback_data="education_vo")
        )
        markup.add(back_button("back_to_search"))
        bot.edit_message_text("Выберите тип образования:", call.message.chat.id, call.message.message_id, reply_markup=markup)

    elif call.data == "back_to_day_selection":
        # Возврат на выбор дня недели
        markup = types.InlineKeyboardMarkup()
        for day in DAYS_MAPPING.keys():
            markup.add(types.InlineKeyboardButton(day.capitalize(), callback_data=f"day_{day}"))
        
        # Кнопка "Назад" возвращает на выбор типа образования
        markup.add(back_button("back_to_group_selection"))

        bot.edit_message_text("Выберите день недели:", call.message.chat.id, call.message.message_id, reply_markup=markup)


ADMIN_ID = 6328346430  # Замените на ваш Telegram ID

@bot.message_handler(content_types=['document'])
def handle_document(message):
    """Обработчик загрузки расписания (только для администратора)"""
    chat_id = message.chat.id

    if chat_id != ADMIN_ID:
        bot.send_message(chat_id, "❌ У вас нет прав загружать расписание.")
        return

    # Сохраняем ID файла
    user_data[chat_id] = {"file_id": message.document.file_id}

    # Запрашиваем, для какого типа обучения загружать файл
    markup = types.InlineKeyboardMarkup()
    markup.add(
        types.InlineKeyboardButton("📚 СПО", callback_data="schedule_spo"),
        types.InlineKeyboardButton("🏛 ВО", callback_data="schedule_vo")
    )

    bot.send_message(chat_id, "📂 Выберите тип обучения:", reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data in ["schedule_spo", "schedule_vo"])
def choose_week_upload(call):
    """Выбор, обновить текущее расписание или загрузить на следующую неделю"""
    chat_id = call.message.chat.id

    # Сохраняем тип расписания (СПО или ВО)
    user_data[chat_id]["schedule_type"] = call.data.split("_")[1]

    markup = types.InlineKeyboardMarkup()
    markup.add(
        types.InlineKeyboardButton("📅 Обновить текущее расописание", callback_data="update_this"),
        types.InlineKeyboardButton("⏭ Расписание на следующую неделю", callback_data="update_next")
    )

    bot.send_message(chat_id, "📂 Куда загрузить расписание?", reply_markup=markup)



def auto_update_schedule():
    """Автоматически обновляет расписание в понедельник"""
    while True:
        now = datetime.now()
        if now.weekday() == 6:  # 0 - это понедельник
            try:
                # Удаляем старые файлы this_spo.xlsx и this_vo.xlsx
                if os.path.exists("this_spo.xlsx"):
                    os.remove("this_spo.xlsx")
                if os.path.exists("this_vo.xlsx"):
                    os.remove("this_vo.xlsx")

                # Переименовываем next в this
                if os.path.exists("next_spo.xlsx"):
                    os.rename("next_spo.xlsx", "this_spo.xlsx")
                if os.path.exists("next_vo.xlsx"):
                    os.rename("next_vo.xlsx", "this_vo.xlsx")

                print("✅ Автоматическое обновление расписания выполнено.")
            except Exception as e:
                print(f"❌ Ошибка при обновлении расписания: {e}")

            # Ждем 24 часа, чтобы избежать повторного срабатывания в один день
            time.sleep(86400)
        else:
            # Проверяем раз в 6 часов
            time.sleep(21600)

# Запускаем автоматическое обновление в отдельном потоке
update_thread = threading.Thread(target=auto_update_schedule, daemon=True)
update_thread.start()

USER_DB_FILE = "users.json"

def load_users():
    """Загружает список пользователей из файла"""
    try:
        with open(USER_DB_FILE, "r") as f:
            return json.load(f)
    except FileNotFoundError:
        return {}

def save_users(users):
    """Сохраняет список пользователей в файл"""
    with open(USER_DB_FILE, "w") as f:
        json.dump(users, f)

users = load_users()

def notify_users():
    """Отправляет сообщение всем пользователям о новом расписании"""
    text = "❗Внимание, студенты❗ Доступно новое расписание СПО❗ Доступно новое расписание ВО❗"
    
    for user_id in users.keys():
        try:
            bot.send_message(user_id, text)
        except Exception as e:
            print(f"Не удалось отправить сообщение пользователю {user_id}: {e}")


@bot.callback_query_handler(func=lambda call: call.data in ["update_this", "update_next"])
def process_schedule_upload(call):
    """Обрабатывает выбор загрузки расписания"""
    chat_id = call.message.chat.id

    if chat_id != ADMIN_ID:
        bot.send_message(chat_id, "❌ У вас нет прав загружать расписание.")
        return

    # Проверяем, есть ли сохраненный тип расписания (СПО или ВО)
    schedule_type = user_data.get(chat_id, {}).get("schedule_type")
    if not schedule_type:
        bot.send_message(chat_id, "❌ Ошибка: Не указан тип расписания (СПО или ВО).")
        return

    # Получаем ID файла
    file_id = user_data.get(chat_id, {}).get("file_id")
    if not file_id:
        bot.send_message(chat_id, "❌ Файл не найден. Пожалуйста, загрузите его заново.")
        return

    # Получаем сам файл
    file_info = bot.get_file(file_id)
    downloaded_file = bot.download_file(file_info.file_path)

    # Определяем имя файла с учетом выбора СПО / ВО
    if call.data == "update_this":
        file_name = f"this_{schedule_type}.xlsx"  # Например, this_spo.xlsx или this_vo.xlsx
    else:
        file_name = f"next_{schedule_type}.xlsx"  # Например, next_spo.xlsx или next_vo.xlsx

    # Сохраняем файл
    with open(file_name, 'wb') as new_file:
        new_file.write(downloaded_file)

    bot.send_message(chat_id, f"✅ Файл сохранен как `{file_name}`.")

    # Уведомляем всех пользователей
    notify_users()

    # Очищаем user_data
    user_data.pop(chat_id, None)

bot.infinity_polling()