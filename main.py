import openpyxl
import telebot
from telebot import *
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton
from telebot import types
import os
import threading
import time
from datetime import datetime
import json
from get_exel_data import *
# Маппинг дней недели и ячеек для времени и пар
def load_mapping(file_name="mapping.json"):
    with open(file_name, "r", encoding="utf-8") as f:
        return json.load(f)

mapping = load_mapping()
DAYS_MAPPING = mapping["DAYS_MAPPING"]
VO_DAYS_MAPPING = mapping["VO_DAYS_MAPPING"]
GROUP_ROOM_MAPPING = mapping["GROUP_ROOM_MAPPING"]

token = ""
bot = telebot.TeleBot(token)
user_data = {}
def back_button(callback_data):
    """Создает кнопку 'Назад' для возврата на предыдущий этап."""
    return types.InlineKeyboardButton("🔙 Назад", callback_data=callback_data)
# Обработчик команды /start
@bot.message_handler(commands=['start', 'help'])
def send_welcome(message):
    chat_id = message.chat.id
    
    # Добавляем пользователя в список, если его нет
    if str(chat_id) not in users:
        users[str(chat_id)] = True
        save_users(users)
    markup = types.InlineKeyboardMarkup()
    markup.add(
        types.InlineKeyboardButton("Студент", callback_data="search_student"),
        types.InlineKeyboardButton("Преподаватель", callback_data="search_teacher")
    )
    bot.send_message(message.chat.id, "Выберите, кого искать:", reply_markup=markup)

# Обработчик выбора "Студент"
@bot.callback_query_handler(func=lambda call: call.data == "search_student")
def select_education_type(call):
    chat_id = call.message.chat.id

    # Полностью очищаем данные пользователя перед новым поиском
    if chat_id in user_data:
        user_data.pop(chat_id)  # Удаляем все старые данные
    user_data[chat_id] = {"search_type": "student"}  # Устанавливаем новый контекст поиска

    # Удаляем предыдущее сообщение
    try:
        bot.delete_message(chat_id, call.message.message_id)
    except Exception as e:
        print(f"Не удалось удалить сообщение: {e}")

    # Отправляем сообщение с выбором типа образования
    markup = types.InlineKeyboardMarkup()
    markup.add(
        types.InlineKeyboardButton("SPO", callback_data="education_spo"),
        types.InlineKeyboardButton("VO", callback_data="education_vo")
    )
    markup.add(back_button("back_to_search"))  # Кнопка "Назад"

    sent_message = bot.send_message(chat_id, "Выберите тип образования:", reply_markup=markup)

    # Сохраняем ID последнего сообщения
    user_data[chat_id]["last_bot_message_id"] = sent_message.message_id


# Обработчик выбора недели
@bot.callback_query_handler(func=lambda call: call.data.startswith("week_"))
def select_week(call):
    week = call.data.split("_")[1]
    user_data[call.message.chat.id]["week"] = week  # Сохраняем выбранную неделю

    # Удаляем старое сообщение
    try:
        bot.delete_message(call.message.chat.id, call.message.message_id)
    except Exception as e:
        print(f"Не удалось удалить сообщение: {e}")

    # Переходим к этапу выбора дня недели
    markup = types.InlineKeyboardMarkup()
    for day in DAYS_MAPPING.keys():
        markup.add(types.InlineKeyboardButton(day.capitalize(), callback_data=f"day_{day}"))
    markup.add(back_button("education_spo"))  # Кнопка "Назад" на предыдущий этап
    bot.send_message(call.message.chat.id, "Выберите день недели:", reply_markup=markup)

# Модификация функции выбора дня недели
def select_day(message):
    user_data[message.chat.id]["group_name"] = message.text.strip()

    # Удаляем сообщение пользователя
    try:
        bot.delete_message(message.chat.id, message.message_id)
    except Exception as e:
        print(f"Не удалось удалить сообщение пользователя: {e}")

    # Удаляем предыдущее сообщение бота
    last_bot_message_id = user_data[message.chat.id].get("last_bot_message_id")
    if last_bot_message_id:
        try:
            bot.delete_message(message.chat.id, last_bot_message_id)
        except Exception as e:
            print(f"Не удалось удалить сообщение бота: {e}")

    # Отправляем сообщение с выбором недели
    markup = types.InlineKeyboardMarkup()
    markup.add(
        types.InlineKeyboardButton("Эта неделя", callback_data="week_this"),
        types.InlineKeyboardButton("Следующая неделя", callback_data="week_next")
    )
    markup.add(back_button("education_spo"))  # Кнопка "Назад" на выбор типа образования
    bot.send_message(message.chat.id, "Выберите неделю:", reply_markup=markup)

# Обработчик выбора типа образования для студентаа
@bot.callback_query_handler(func=lambda call: call.data.startswith("education_"))
def ask_group_name(call):
    education_type = call.data.split("_")[1].upper()
    user_data[call.message.chat.id]["education_type"] = education_type

    # Удаляем старое сообщение
    try:
        bot.delete_message(call.message.chat.id, call.message.message_id)
    except Exception as e:
        print(f"Не удалось удалить сообщение: {e}")

    # Очищаем старые обработчики сообщений для этого пользователя
    bot.clear_step_handler_by_chat_id(call.message.chat.id)

    # Отправляем новое сообщение
    markup = types.InlineKeyboardMarkup()
    markup.add(back_button("search_student"))  # Кнопка "Назад"
    sent_message = bot.send_message(call.message.chat.id, "Введите номер группы(например C7124Б. C - eng, Б - ru):", reply_markup=markup)
    
    # Сохраняем ID отправленного сообщения для последующего удаления
    user_data[call.message.chat.id]["last_bot_message_id"] = sent_message.message_id
    bot.register_next_step_handler(sent_message, select_day)

# Обработчик выбора дня недели для студента
@bot.callback_query_handler(func=lambda call: call.data.startswith("day_"))
def show_schedule(call):
    day = call.data.split("_")[1]
    user_data[call.message.chat.id]["day"] = day
    week = user_data[call.message.chat.id].get("week")
    education_type = user_data[call.message.chat.id]["education_type"]
    group_name = user_data[call.message.chat.id]["group_name"]

    # Определяем файл в зависимости от недели и типа образования
    if week == "this":
        file_name = "this_spo.xlsx" if education_type == "SPO" else "this_vo.xlsx"
    elif week == "next":
        file_name = "next_spo.xlsx" if education_type == "SPO" else "next_vo.xlsx"

    try:
        workbook = openpyxl.load_workbook(file_name, data_only=True)
    except FileNotFoundError:
        bot.send_message(call.message.chat.id, f"Ошибка: Файл {file_name} не найден.")
        return

    # Оставшаяся часть логики аналогична
    found_group = False
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        column = get_group_data(sheet, group_name)
        if column:
            schedule = extract_schedule(sheet, column, day, education_type, sheet_name)
            
            # Удаление старого сообщения
            try:
                bot.delete_message(call.message.chat.id, call.message.message_id)
            except Exception as e:
                print(f"Не удалось удалить сообщение: {e}")

            # Если расписание пустое, отправляем сообщение об отсутствии пар
            if not schedule:
                markup = types.InlineKeyboardMarkup()
                markup.add(back_button("back_to_day_selection"))  # Кнопка "Назад" на выбор дня недели
                bot.send_message(call.message.chat.id, f"На {day.capitalize()} для группы {group_name} пар нет.", reply_markup=markup)

            else:
                markup = types.InlineKeyboardMarkup()
                markup.add(back_button("back_to_day_selection"))  # Кнопка возвращает на выбор дня недели
                bot.send_message(call.message.chat.id, display_schedule(schedule, group_name), reply_markup=markup)

            found_group = True
            break

    if not found_group:
        bot.send_message(call.message.chat.id, f"Группа '{group_name}' не найдена в файле {file_name}.")

# Обработчик выбора "Преподаватель"
@bot.callback_query_handler(func=lambda call: call.data == "search_teacher")
def ask_teacher_name(call):
    user_data[call.message.chat.id] = {"search_type": "teacher"}

    # Удаляем предыдущее сообщение
    try:
        bot.delete_message(call.message.chat.id, call.message.message_id)
    except Exception as e:
        print(f"Не удалось удалить сообщение: {e}")

    # Очищаем все старые обработчики сообщений
    bot.clear_step_handler_by_chat_id(call.message.chat.id)

    # Создаем клавиатуру с кнопкой "Назад"
    markup = types.InlineKeyboardMarkup()
    markup.add(back_button("back_to_search"))

    # Отправляем сообщение с запросом фамилии и инициалов преподавателя
    sent_message = bot.send_message(call.message.chat.id, "Введите фамилию и инициалы преподавателя (например, Иванов И. И.):", reply_markup=markup)

    # Сохраняем ID отправленного сообщения для последующего удаления
    user_data[call.message.chat.id]["last_bot_message_id"] = sent_message.message_id
    bot.register_next_step_handler(sent_message, select_teacher_week_step)


def select_teacher_week_step(message):
    user_data[message.chat.id]["teacher_name"] = message.text.strip()

    # Удаляем сообщение пользователя
    try:
        bot.delete_message(message.chat.id, message.message_id)
    except Exception as e:
        print(f"Не удалось удалить сообщение пользователя: {e}")

    # Удаляем сообщение "Введите фамилию и инициалы преподавателя"
    last_bot_message_id = user_data[message.chat.id].get("last_bot_message_id")
    if last_bot_message_id:
        try:
            bot.delete_message(message.chat.id, last_bot_message_id)
        except Exception as e:
            print(f"Не удалось удалить сообщение бота: {e}")

    # Отправляем сообщение с выбором недели
    markup = types.InlineKeyboardMarkup()
    markup.add(
        types.InlineKeyboardButton("Эта неделя", callback_data="teacher_week_this"),
        types.InlineKeyboardButton("Следующая неделя", callback_data="teacher_week_next")
    )
    markup.add(back_button("search_teacher"))

    # Отправляем сообщение с выбором недели и сохраняем его ID
    sent_message = bot.send_message(message.chat.id, "Выберите неделю для преподавателя:", reply_markup=markup)
    user_data[message.chat.id]["last_bot_message_id"] = sent_message.message_id

@bot.callback_query_handler(func=lambda call: call.data.startswith("teacher_week_"))
def select_teacher_week(call):
    """Обрабатывает выбор недели преподавателем."""
    week = call.data.split("_")[2]
    user_data[call.message.chat.id]["week"] = week  # Сохраняем выбранную неделю

    # Удаляем старое сообщение (с выбором недели)
    try:
        bot.delete_message(call.message.chat.id, call.message.message_id)
    except Exception as e:
        print(f"Не удалось удалить сообщение: {e}")

    # Создаем клавиатуру
    markup = InlineKeyboardMarkup(row_width=2)  # Устанавливаем 2 кнопки в ряд

    # Добавляем кнопку для показа расписания на всю неделю (отдельным рядом)
    markup.add(InlineKeyboardButton("📅 Показать расписание на всю неделю", callback_data="show_week_schedule"))

    # Добавляем кнопки для дней недели (в два ряда)
    day_buttons = [InlineKeyboardButton(day.capitalize(), callback_data=f"teacher_day_{day}") for day in DAYS_MAPPING.keys()]
    markup.add(*day_buttons)

    # Добавляем кнопку "Назад" для возврата на предыдущий этап
    markup.add(back_button("search_teacher"))

    # Отправляем сообщение с клавиатурой
    bot.send_message(call.message.chat.id, "Выберите день недели или просмотрите расписание на всю неделю:", reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data == "show_week_schedule")
def show_week_schedule(call):
    """Обработчик для показа полного расписания преподавателя на выбранную неделю."""
    chat_id = call.message.chat.id
    teacher_name = user_data[chat_id].get("teacher_name")
    week = user_data[chat_id].get("week")

    if not teacher_name or not week:
        bot.send_message(chat_id, "Ошибка: Не выбрано имя преподавателя или неделя.")
        return

    # Загружаем расписание на неделю
    try:
        workbook = openpyxl.load_workbook(f"{week}_spo.xlsx", data_only=True)
    except FileNotFoundError:
        bot.send_message(chat_id, f"Ошибка: файл расписания на {week} не найден.")
        return

    schedule = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for day in DAYS_MAPPING.keys():
            daily_schedule = search_teacher(sheet, sheet_name, teacher_name, target_day=day, education_type="SPO")
            schedule.extend(daily_schedule)

    if not schedule:
        bot.send_message(chat_id, f"На неделю для преподавателя {teacher_name} расписание отсутствует.")
        return

    # Генерация изображения расписания
    image_path = generate_week_schedule_image(schedule, teacher_name)

    # Отправляем изображение
    with open(image_path, "rb") as img:
        bot.send_document(chat_id, img)

    # Удаляем изображение с сервера
    if os.path.exists(image_path):
        os.remove(image_path)

@bot.callback_query_handler(func=lambda call: call.data.startswith("teacher_day_"))
def show_teacher_schedule(call):
    """ Показывает расписание преподавателя """
    day = call.data.split("_")[2]
    chat_id = call.message.chat.id
    teacher_name = user_data[chat_id]["teacher_name"]
    week = user_data[chat_id].get("week")

    # Файлы для текущей или следующей недели
    file_names = ["this_spo.xlsx", "this_vo.xlsx"] if week == "this" else ["next_spo.xlsx", "next_vo.xlsx"]

    print(f"[DEBUG] Поиск для {teacher_name}, день: {day}, неделя: {week}")

    found_data = []

    for file_name in file_names:
        try:
            workbook = openpyxl.load_workbook(file_name, data_only=True)
        except FileNotFoundError:
            bot.send_message(chat_id, f"Ошибка: Файл {file_name} не найден.")
            continue

        education_type = "SPO" if "spo" in file_name.lower() else "VO"

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Для VO проверяем, есть ли маппинг
            if education_type == "VO":
                vo_mapping = determine_vo_sheet_mapping(sheet_name)
                if not vo_mapping:
                    continue

            # Поиск расписания преподавателя
            results = search_teacher(sheet, sheet_name, teacher_name, target_day=day, education_type=education_type)
            found_data.extend(results)

    # **Добавляем сортировку по номеру пары и времени**
    found_data.sort(key=lambda x: (
        int(x.get("pair", "0")) if x.get("pair", "0").isdigit() else 0,  # Сортируем по номеру пары
        x.get("time", ["00:00"])[0]  # Если пары одинаковые, сортируем по началу занятия
    ))

    # Удаление старого сообщения
    try:
        bot.delete_message(chat_id, call.message.message_id)
    except Exception as e:
        print(f"Ошибка удаления сообщения: {e}")

    # Вывод результата
    markup = types.InlineKeyboardMarkup()
    markup.add(back_button(f"teacher_week_{week}"))

    if not found_data:
        bot.send_message(chat_id, f"На {day.capitalize()} для преподавателя {teacher_name} пар нет.", reply_markup=markup)
    else:
        bot.send_message(chat_id, display_schedule(found_data, teacher_name), reply_markup=markup)

# Обработчик кнопки "Назад"
@bot.callback_query_handler(func=lambda call: call.data.startswith("back_"))
def go_back(call):
    chat_id = call.message.chat.id

    # Удаляем текущее сообщение
    try:
        bot.delete_message(chat_id, call.message.message_id)
    except Exception as e:
        print(f"Ошибка удаления сообщения: {e}")

    # Определяем текущий контекст поиска (студент или преподаватель)
    search_type = user_data.get(chat_id, {}).get("search_type")

    if call.data == "back_to_search":
        # Возврат к выбору "Студент/Преподаватель"
        markup = types.InlineKeyboardMarkup()
        markup.add(
            types.InlineKeyboardButton("Студент", callback_data="search_student"),
            types.InlineKeyboardButton("Преподаватель", callback_data="search_teacher")
        )
        bot.send_message(chat_id, "Выберите, кого искать:", reply_markup=markup)

    elif call.data == "back_to_group_selection" and search_type == "student":
        # Возврат на выбор типа образования для студента
        markup = types.InlineKeyboardMarkup()
        markup.add(
            types.InlineKeyboardButton("SPO", callback_data="education_spo"),
            types.InlineKeyboardButton("VO", callback_data="education_vo")
        )
        markup.add(back_button("back_to_search"))  # Кнопка "Назад" ведёт к главному меню
        bot.send_message(chat_id, "Выберите тип образования:", reply_markup=markup)

    elif call.data == "back_to_teacher_search" and search_type == "teacher":
        # Возврат к запросу фамилии преподавателя
        markup = types.InlineKeyboardMarkup()
        markup.add(back_button("back_to_search"))
        sent_message = bot.send_message(chat_id, "Введите фамилию и инициалы преподавателя (например, Иванов И. И.):", reply_markup=markup)
        
        # Сохраняем ID сообщения, чтобы удалить его при следующем вызове
        user_data[chat_id]["last_bot_message_id"] = sent_message.message_id
        bot.register_next_step_handler(sent_message, select_teacher_week_step)

    elif call.data == "back_to_day_selection" and search_type == "student":
        # Возврат на выбор дня недели для студента
        markup = InlineKeyboardMarkup()
        for day in DAYS_MAPPING.keys():
            markup.add(InlineKeyboardButton(day.capitalize(), callback_data=f"day_{day}"))
        markup.add(back_button("back_to_group_selection"))  # Назад на выбор группы
        bot.send_message(chat_id, "Выберите день недели:", reply_markup=markup)

    elif call.data == "back_to_day_selection" and search_type == "teacher":
        # Возвращаем преподавателя к выбору дня недели
        markup = InlineKeyboardMarkup()
        for day in DAYS_MAPPING.keys():
            markup.add(InlineKeyboardButton(day.capitalize(), callback_data=f"teacher_day_{day}"))
        markup.add(back_button("back_to_teacher_search"))  # Назад на выбор преподавателя
        bot.send_message(chat_id, "Выберите день недели:", reply_markup=markup)

    elif call.data == "back_to_week_selection" and search_type == "student":
        # Возвращение к выбору недели для студента
        markup = types.InlineKeyboardMarkup()
        markup.add(
            types.InlineKeyboardButton("Эта неделя", callback_data="week_this"),
            types.InlineKeyboardButton("Следующая неделя", callback_data="week_next")
        )
        markup.add(back_button("back_to_group_selection"))  # Назад на выбор группы
        bot.send_message(chat_id, "Выберите неделю:", reply_markup=markup)

    elif call.data == "back_to_week_selection" and search_type == "teacher":
        # Возвращение к выбору недели для преподавателя
        markup = types.InlineKeyboardMarkup()
        markup.add(
            types.InlineKeyboardButton("Эта неделя", callback_data="teacher_week_this"),
            types.InlineKeyboardButton("Следующая неделя", callback_data="teacher_week_next")
        )
        markup.add(back_button("back_to_teacher_search"))
        bot.send_message(chat_id, "Выберите неделю для преподавателя:", reply_markup=markup)


ADMIN_ID = 6328346430  # Замените на ваш Telegram ID

@bot.message_handler(content_types=['document'])
def handle_document(message):
    """Обработчик загрузки расписания (только для администратора)"""
    chat_id = message.chat.id

    if chat_id != ADMIN_ID:
        bot.send_message(chat_id, "❌ У вас нет прав загружать расписание.")
        return

    # Сохраняем ID файла
    user_data[chat_id] = {"file_id": message.document.file_id}

    # Запрашиваем, для какого типа обучения загружать файл
    markup = types.InlineKeyboardMarkup()
    markup.add(
        types.InlineKeyboardButton("📚 СПО", callback_data="schedule_spo"),
        types.InlineKeyboardButton("🏛 ВО", callback_data="schedule_vo")
    )

    bot.send_message(chat_id, "📂 Выберите тип обучения:", reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data in ["schedule_spo", "schedule_vo"])
def choose_week_upload(call):
    """Выбор, обновить текущее расписание или загрузить на следующую неделю"""
    chat_id = call.message.chat.id

    # Сохраняем тип расписания (СПО или ВО)
    user_data[chat_id]["schedule_type"] = call.data.split("_")[1]

    markup = types.InlineKeyboardMarkup()
    markup.add(
        types.InlineKeyboardButton("📅 Обновить текущее расописание", callback_data="update_this"),
        types.InlineKeyboardButton("⏭ Расписание на следующую неделю", callback_data="update_next")
    )

    bot.send_message(chat_id, "📂 Куда загрузить расписание?", reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data in ["update_this", "update_next"])
def process_schedule_upload(call):
    """Обрабатывает выбор загрузки расписания"""
    chat_id = call.message.chat.id

    if chat_id != ADMIN_ID:
        bot.send_message(chat_id, "❌ У вас нет прав загружать расписание.")
        return

    # Проверяем, есть ли сохраненный тип расписания (СПО или ВО)
    schedule_type = user_data.get(chat_id, {}).get("schedule_type")
    if not schedule_type:
        bot.send_message(chat_id, "❌ Ошибка: Не указан тип расписания (СПО или ВО).")
        return

    # Получаем ID файла
    file_id = user_data.get(chat_id, {}).get("file_id")
    if not file_id:
        bot.send_message(chat_id, "❌ Файл не найден. Пожалуйста, загрузите его заново.")
        return

    # Получаем сам файл
    file_info = bot.get_file(file_id)
    downloaded_file = bot.download_file(file_info.file_path)

    # Определяем имя файла с учетом выбора СПО / ВО
    if call.data == "update_this":
        file_name = f"this_{schedule_type}.xlsx"  # Например, this_spo.xlsx или this_vo.xlsx
    else:
        file_name = f"next_{schedule_type}.xlsx"  # Например, next_spo.xlsx или next_vo.xlsx

    # Сохраняем файл
    with open(file_name, 'wb') as new_file:
        new_file.write(downloaded_file)

    bot.send_message(chat_id, f"✅ Файл сохранен как `{file_name}`.")

    # Уведомляем всех пользователей
    notify_users()

    # Очищаем user_data
    user_data.pop(chat_id, None)

def auto_update_schedule():
    """Автоматически обновляет расписание в понедельник"""
    while True:
        now = datetime.now()
        if now.weekday() == 0:  # 0 - это понедельник
            try:
                # Удаляем старые файлы this_spo.xlsx и this_vo.xlsx
                if os.path.exists("this_spo.xlsx"):
                    os.remove("this_spo.xlsx")
                if os.path.exists("this_vo.xlsx"):
                    os.remove("this_vo.xlsx")

                # Переименовываем next в this
                if os.path.exists("next_spo.xlsx"):
                    os.rename("next_spo.xlsx", "this_spo.xlsx")
                if os.path.exists("next_vo.xlsx"):
                    os.rename("next_vo.xlsx", "this_vo.xlsx")

                print("✅ Автоматическое обновление расписания выполнено.")
            except Exception as e:
                print(f"❌ Ошибка при обновлении расписания: {e}")

            # Ждем 24 часа, чтобы избежать повторного срабатывания в один день
            time.sleep(86400)
        else:
            # Проверяем раз в 6 часов
            time.sleep(21600)

# Запускаем автоматическое обновление в отдельном потоке
update_thread = threading.Thread(target=auto_update_schedule, daemon=True)
update_thread.start()

USER_DB_FILE = "users.json"

def load_users():
    """Загружает список пользователей из файла"""
    try:
        with open(USER_DB_FILE, "r") as f:
            return json.load(f)
    except FileNotFoundError:
        return {}

def save_users(users):
    """Сохраняет список пользователей в файл"""
    with open(USER_DB_FILE, "w") as f:
        json.dump(users, f)

users = load_users()

def notify_users():
    """Отправляет сообщение всем пользователям о новом расписании"""
    text = "❗Внимание, студенты❗ Доступно новое расписание СПО❗ Доступно новое расписание ВО❗"
    
    for user_id in users.keys():
        try:
            bot.send_message(user_id, text)
        except Exception as e:
            print(f"Не удалось отправить сообщение пользователю {user_id}: {e}")

bot.infinity_polling()